"""
convert_databases.py
Reads the three .xlsx source files and writes compact JSON data files
consumed by the ChemE Units web app.

Run once:  python convert_databases.py
"""

import json
import os
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE, "data")
os.makedirs(DATA_DIR, exist_ok=True)

def safe(v):
    """Return a JSON-safe value (handle None, floats, etc.)."""
    if v is None:
        return None
    if isinstance(v, float):
        return v
    if isinstance(v, int):
        return v
    return str(v)

# ─────────────────────────────────────────────────────────────────────────────
# 1. UNITS DATABASE
# ─────────────────────────────────────────────────────────────────────────────
print("Processing units_database.xlsx …")
wb_u = openpyxl.load_workbook(os.path.join(BASE, "units_database.xlsx"))

# --- units (main sheet) ---
ws_u = wb_u["units_database"]
units = {}
headers = [c.value for c in ws_u[1]]
for row in ws_u.iter_rows(min_row=2, values_only=True):
    sym = row[0]
    factor = row[4]
    if sym and factor is not None and isinstance(factor, (int, float)):
        dim = {
            "M":  row[7]  or 0,
            "L":  row[8]  or 0,
            "T":  row[9]  or 0,
            "Th": row[10] or 0,
            "N":  row[11] or 0,
            "I":  row[12] or 0,
            "J":  row[13] or 0,
        }
        units[str(sym)] = {
            "name":   str(row[1]) if row[1] else str(sym),
            "factor": float(factor),
            "dim":    dim,
        }

# --- SI prefixes ---
ws_p = wb_u["SI_prefixes"]
prefixes = {}
for row in ws_p.iter_rows(min_row=2, values_only=True):
    if row[1] and row[2] is not None:
        prefixes[str(row[1])] = {
            "name":   str(row[0]),
            "factor": float(row[2]),
        }

# --- derived units ---
ws_d = wb_u["derived_units"]
derived = {}
for row in ws_d.iter_rows(min_row=2, values_only=True):
    sym = row[0]
    if not sym:
        continue
    dim = {
        "M":  row[3] or 0,
        "L":  row[4] or 0,
        "T":  row[5] or 0,
        "Th": row[6] or 0,
        "N":  row[7] or 0,
        "I":  row[8] or 0,
        "J":  row[9] or 0,
    }
    derived[str(sym)] = {
        "name":   str(row[1]) if row[1] else str(sym),
        "expr":   str(row[2]) if row[2] else "",
        "dim":    dim,
        "factor": float(row[10]) if row[10] is not None else 1.0,
    }

units_data = {
    "units": units,
    "prefixes": prefixes,
    "derived": derived,
}

path_u = os.path.join(DATA_DIR, "units_data.json")
with open(path_u, "w", encoding="utf-8") as f:
    json.dump(units_data, f, ensure_ascii=False, separators=(",", ":"))
print(f"  -> {len(units)} units, {len(prefixes)} prefixes, {len(derived)} derived  ->  {path_u}")

# ─────────────────────────────────────────────────────────────────────────────
# 2. PERIODIC TABLE
# ─────────────────────────────────────────────────────────────────────────────
print("Processing periodic_table.xlsx …")
wb_pt = openpyxl.load_workbook(os.path.join(BASE, "periodic_table.xlsx"))
ws_pt = wb_pt["periodic_table"]

elements = []
for row in ws_pt.iter_rows(min_row=2, values_only=True):
    Z = row[0]
    sym = row[1]
    name = row[2]
    aw = row[3]
    if Z is None or sym is None:
        continue
    elements.append({
        "Z":       int(Z),
        "symbol":  str(sym),
        "name":    str(name) if name else str(sym),
        "weight":  float(aw) if isinstance(aw, (int, float)) else None,
        "category": str(row[6]) if row[6] else "Unknown",
        "period":  int(row[7]) if row[7] else None,
        "group":   int(row[8]) if isinstance(row[8], (int, float)) else None,
        "block":   str(row[9]) if row[9] else None,
        "phase":   str(row[18]) if row[18] else None,
        "melt_K":  float(row[14]) if isinstance(row[14], (int, float)) else None,
        "boil_K":  float(row[15]) if isinstance(row[15], (int, float)) else None,
        "density": float(row[16]) if isinstance(row[16], (int, float)) else None,
        "electronegativity": float(row[11]) if isinstance(row[11], (int, float)) else None,
    })

path_pt = os.path.join(DATA_DIR, "periodic_table.json")
with open(path_pt, "w", encoding="utf-8") as f:
    json.dump(elements, f, ensure_ascii=False, separators=(",", ":"))
print(f"  -> {len(elements)} elements  ->  {path_pt}")

# ─────────────────────────────────────────────────────────────────────────────
# 3. ENGINEERING CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
print("Processing engineering_constants.xlsx …")
wb_c = openpyxl.load_workbook(os.path.join(BASE, "engineering_constants.xlsx"))
ws_c = wb_c["engineering_constants"]

constants = []
current_category = "General"
for row in ws_c.iter_rows(min_row=2, values_only=True):
    sym = row[0]
    name = row[1]
    value = row[2]
    unit = row[4]
    cat = row[7]

    # Category header rows (only first col has value, rest are None)
    if sym and value is None and name is None:
        current_category = str(sym)
        continue
    if sym and name and isinstance(value, (int, float)):
        constants.append({
            "symbol":   str(sym),
            "name":     str(name),
            "value":    float(value),
            "unit":     str(unit) if unit else "",
            "category": str(cat) if cat else current_category,
            "uncertainty": str(row[3]) if row[3] else "",
        })

path_c = os.path.join(DATA_DIR, "constants.json")
with open(path_c, "w", encoding="utf-8") as f:
    json.dump(constants, f, ensure_ascii=False, separators=(",", ":"))
print(f"  -> {len(constants)} constants  ->  {path_c}")

print("\nAll done! JSON files are in ./data/")
